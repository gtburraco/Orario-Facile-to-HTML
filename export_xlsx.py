from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Side, Border
from modelli.docente import Docente
import config

def esporta_xlsx(docenti: List[Docente], file_path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Orario Docenti"
    ws.merge_cells("A1:A2")
    ws["A1"] = "Docente"

    bordo_spesso = Border(left = Side(style="medium"))

    i = 3
    for d in docenti:
        cell = ws.cell(row=i, column=1, value=d.nome)
        i += 1

    max_len = 0
    for cell in ws["A"]:
        if cell.value:
            max_len = max(max_len, len(str(cell.value)))
    ws.column_dimensions["A"].width = max_len + 1


    col = 2
    for giorno in config.giorni_di_scuola:
        ora = 1
        fasce_per_questo_giorno = max(f.blocco_orario for f in config.fasce_settimanali[giorno])
        for blocco in range(fasce_per_questo_giorno):
            ws.cell(row=1, column=col, value=giorno)
            cell = ws.cell(row=2, column=col, value=ora)
            cell.border = Border(bottom = Side(style="medium"))
            col += 1
            ora += 1

    start_riga_docente = 3
    start_colonna_giorno = 2
    for docente in docenti:
        for giorno in config.giorni_di_scuola:

            cell = ws.cell(row=start_riga_docente, column=start_colonna_giorno)
            cell.border = bordo_spesso

            fasce_per_questo_giorno = max(f.blocco_orario for f in config.fasce_settimanali[giorno])
            for i in range(fasce_per_questo_giorno):
                blocco_attuale = i+1
                risultati = [l for l in docente.lezioni if l.fascia_oraria.blocco_orario==blocco_attuale and l.giorno==giorno]
                if risultati:
                    risultato_stringa = "\n".join([s.get_classi_scolastiche_xlsx() for s in risultati])
                    cell = ws.cell(row=start_riga_docente,column=start_colonna_giorno,value=risultato_stringa)
                    cell.alignment = Alignment(wrapText=True, horizontal="center", vertical="center")
                start_colonna_giorno   +=1

        start_riga_docente += 1
        start_colonna_giorno = 2

    for cell in ws["A"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="left", vertical="center")

    for cell in ws["1"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for cell in ws["2"]:
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")

    ws.freeze_panes = "B3"
    ws.print_title_rows = "1:2"
    ws.print_title_cols = "A:A"
    wb.save(file_path)

